"""
Automatizacion del Procesamiento de Resultados COLDEX Sectorial.
Genera un Excel con una hoja por cada combinacion Sector x Tipo (IND/CAD).

Pure Python + openpyxl (sin pandas/numpy) para mantener el deploy
serverless por debajo del limite de tamano de Vercel.
"""

import io
import math
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ─── Configuracion ──────────────────────────────────────────────────────────

DATA_DIR = Path("data")
OUTPUT_FILE = DATA_DIR / "Procesamiento_Resultados_Coldex_Sectorial_2025.xlsx"

SECTOR_MAP = {
    "TXT": "Textil",
    "ELT": "Electro",
    "CNS": "Consumo",
    "HGR": "Hogar",
    "SLD": "Salud",
    "SLDI": "Salud Insumos",
}

TYPE_MAP = {
    "INDUSTRIAL": "IND",
    "CADENA": "CAD",
}

SUBCATEGORY_ORDER = [
    "Gestión Comercial",
    "Relacionamiento y Comunicación",
    "Calidad de Datos",
    "Estándares",
    "Gestión de Devoluciones",
    "Gestión de Pedidos",
    "Gestión Logística",
    "Operadores Logísticos",
    "Omnicanalidad",
    "Gestión de Indicadores",
    "Gestión de la Demanda",
    "Gestión de Punto de Venta",
    "Sostenibilidad",
]


# ─── Utilidades de datos (reemplazan pandas) ────────────────────────────────

def is_num(v):
    if v is None:
        return False
    if isinstance(v, bool):
        return False
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return False
        return True
    return False


def read_xlsx_records(source):
    """Lee un xlsx (path o bytes) y devuelve lista de dicts."""
    if isinstance(source, (bytes, bytearray)):
        source = io.BytesIO(source)
    wb = load_workbook(source, read_only=True, data_only=True)
    ws = wb.active
    it = ws.iter_rows(values_only=True)
    headers = [str(h) if h is not None else "" for h in next(it)]
    rows = []
    for row in it:
        if all(v is None for v in row):
            continue
        rows.append(dict(zip(headers, row)))
    wb.close()
    return rows


def filter_rows(rows, **kwargs):
    return [r for r in rows if all(r.get(k) == v for k, v in kwargs.items())]


def unique_values(rows, col):
    return {r.get(col) for r in rows if r.get(col) is not None}


def unique_sorted(rows, col):
    return sorted(unique_values(rows, col))


def nunique(rows, col):
    return len(unique_values(rows, col))


def mean(values):
    nums = [v for v in values if is_num(v)]
    return sum(nums) / len(nums) if nums else None


def mean_col(rows, col):
    return mean([r.get(col) for r in rows])


def round2(v):
    return round(v, 2) if v is not None else None


def group_nunique(rows, by_col, count_col):
    """Cuenta valores unicos de count_col por cada grupo de by_col."""
    d = {}
    for r in rows:
        key = r.get(by_col)
        if key is None:
            continue
        val = r.get(count_col)
        if val is None:
            continue
        d.setdefault(key, set()).add(val)
    return {k: len(v) for k, v in d.items()}


def pivot_mean(rows, row_key_cols, col_key, val_col):
    """Pivot: media de val_col agrupada por tuple(row_key_cols) x col_key.

    Devuelve dict {row_tuple: {col_value: mean}}.
    """
    buckets = {}
    for r in rows:
        rk = tuple(r.get(k) for k in row_key_cols)
        ck = r.get(col_key)
        val = r.get(val_col)
        if not is_num(val):
            continue
        buckets.setdefault(rk, {}).setdefault(ck, []).append(val)
    return {rk: {ck: sum(v) / len(v) for ck, v in inner.items()} for rk, inner in buckets.items()}


def ordered_subcategories(rows):
    present = set(unique_values(rows, "PollDesc2"))
    ordered = [s for s in SUBCATEGORY_ORDER if s in present]
    extras = sorted(present - set(ordered))
    return ordered + extras


def find_input_file():
    xlsx_files = [
        f for f in sorted(DATA_DIR.glob("*.xlsx"))
        if f != OUTPUT_FILE and not f.name.startswith("~$")
    ]
    if not xlsx_files:
        raise FileNotFoundError(f"No se encontro ningun archivo .xlsx en {DATA_DIR}/")
    return xlsx_files[0]


# ─── Estilos ────────────────────────────────────────────────────────────────

FONT_BASE = Font(name="Aptos Narrow", size=11)
FONT_BOLD = Font(name="Aptos Narrow", size=11, bold=True)
FONT_HEADER = Font(name="Aptos Narrow", size=11, bold=True)

FILL_HEADER = PatternFill(start_color="C0E6F5", end_color="C0E6F5", fill_type="solid")
FILL_GREEN = PatternFill(start_color="47D359", end_color="47D359", fill_type="solid")
FILL_CELESTE = PatternFill(start_color="C0E6F5", end_color="C0E6F5", fill_type="solid")

BORDER_THIN_BLUE = Border(
    left=Side(style="thin", color="44B3E1"),
    right=Side(style="thin", color="44B3E1"),
    top=Side(style="thin", color="44B3E1"),
    bottom=Side(style="thin", color="44B3E1"),
)

NUMBER_FORMAT = "0.00"


def set_cell(ws, row, col, value, font=None, fill=None, border=None, number_format=None, alignment=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or FONT_BASE
    if fill:
        cell.fill = fill
    if border:
        cell.border = border
    if number_format and is_num(value):
        cell.number_format = number_format
    if alignment:
        cell.alignment = alignment
    return cell


def apply_row_style(ws, row, max_col, font=None, fill=None, border=None):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        if font:
            cell.font = font
        if fill:
            cell.fill = fill
        if border:
            cell.border = border


# ─── Construccion de hoja ───────────────────────────────────────────────────

def build_sheet(wb, rows_sector, id_poll, id_category, id_type, sector_name, type_short):
    sheet_name = f"Proc {sector_name} {type_short}"[:31]
    ws = wb.create_sheet(title=sheet_name)

    if not rows_sector:
        set_cell(ws, 1, 1, "Sin datos para esta combinacion")
        return

    set_cell(ws, 1, 1, "IDPoll", font=FONT_BOLD)
    set_cell(ws, 1, 2, id_poll)
    set_cell(ws, 2, 1, "IDCategory", font=FONT_BOLD)
    set_cell(ws, 2, 2, id_category)
    set_cell(ws, 3, 1, "IDType", font=FONT_BOLD)
    set_cell(ws, 3, 2, id_type)

    companies = unique_sorted(rows_sector, "CompanyNameTo")
    num_companies = len(companies)
    total_cols = num_companies + 2

    ROW_PIVOT_TITLE = 5
    ROW_PIVOT_HEADER = 6
    ROW_PIVOT_START = 7

    set_cell(ws, ROW_PIVOT_TITLE, 1, "Promedio de Calification", font=FONT_HEADER, fill=FILL_HEADER)
    set_cell(ws, ROW_PIVOT_TITLE, 2, "Etiquetas de columna", font=FONT_HEADER, fill=FILL_HEADER)
    for i in range(3, total_cols + 1):
        set_cell(ws, ROW_PIVOT_TITLE, i, None, fill=FILL_HEADER)

    set_cell(ws, ROW_PIVOT_HEADER, 1, "Etiquetas de fila", font=FONT_HEADER, fill=FILL_HEADER)
    for ci, comp in enumerate(companies):
        set_cell(ws, ROW_PIVOT_HEADER, ci + 2, comp, font=FONT_HEADER, fill=FILL_HEADER)
    set_cell(ws, ROW_PIVOT_HEADER, num_companies + 2, "Total general", font=FONT_HEADER, fill=FILL_HEADER)

    hierarchy_keys = sorted(
        {(
            r["PollLevel1"], r["PollDesc1"],
            r["PollLevel2"], r["PollDesc2"],
            r["PollLevel3"], r["PollDesc3"],
        ) for r in rows_sector},
        key=lambda t: (t[0], t[2], t[4]),
    )

    pivot = pivot_mean(
        rows_sector,
        ["PollLevel1", "PollDesc1", "PollLevel2", "PollDesc2", "PollLevel3", "PollDesc3"],
        "CompanyNameTo",
        "Calification",
    )

    current_row = ROW_PIVOT_START
    prev_level1 = None
    prev_level2 = None

    for (lvl1, desc1, lvl2, desc2, lvl3, desc3) in hierarchy_keys:
        if desc1 != prev_level1:
            set_cell(ws, current_row, 1, desc1, font=FONT_BOLD)
            apply_row_style(ws, current_row, total_cols, border=BORDER_THIN_BLUE)
            ws.cell(row=current_row, column=1).font = FONT_BOLD
            df_l1 = filter_rows(rows_sector, PollDesc1=desc1)
            for ci, comp in enumerate(companies):
                val = mean_col(filter_rows(df_l1, CompanyNameTo=comp), "Calification")
                if val is not None:
                    set_cell(ws, current_row, ci + 2, round2(val), font=FONT_BOLD, number_format=NUMBER_FORMAT)
            val = mean_col(df_l1, "Calification")
            if val is not None:
                set_cell(ws, current_row, num_companies + 2, round2(val), font=FONT_BOLD, number_format=NUMBER_FORMAT)
            current_row += 1
            prev_level1 = desc1
            prev_level2 = None

        if desc2 != prev_level2:
            set_cell(ws, current_row, 1, f"  {desc2}", font=FONT_BOLD)
            apply_row_style(ws, current_row, total_cols, border=BORDER_THIN_BLUE)
            ws.cell(row=current_row, column=1).font = FONT_BOLD
            df_l2 = filter_rows(rows_sector, PollDesc1=desc1, PollDesc2=desc2)
            for ci, comp in enumerate(companies):
                val = mean_col(filter_rows(df_l2, CompanyNameTo=comp), "Calification")
                if val is not None:
                    set_cell(ws, current_row, ci + 2, round2(val), font=FONT_BOLD, number_format=NUMBER_FORMAT)
            val = mean_col(df_l2, "Calification")
            if val is not None:
                set_cell(ws, current_row, num_companies + 2, round2(val), font=FONT_BOLD, number_format=NUMBER_FORMAT)
            current_row += 1
            prev_level2 = desc2

        desc3_display = desc3 if len(str(desc3)) <= 120 else str(desc3)[:117] + "..."
        set_cell(ws, current_row, 1, f"    {desc3_display}")

        key = (lvl1, desc1, lvl2, desc2, lvl3, desc3)
        row_map = pivot.get(key, {})
        company_vals = []
        for ci, comp in enumerate(companies):
            val = row_map.get(comp)
            if is_num(val):
                set_cell(ws, current_row, ci + 2, round2(val), number_format=NUMBER_FORMAT)
                company_vals.append(val)

        if company_vals:
            total_val = sum(company_vals) / len(company_vals)
            set_cell(ws, current_row, num_companies + 2, round2(total_val), number_format=NUMBER_FORMAT)

        current_row += 1

    row_total_general = current_row
    set_cell(ws, row_total_general, 1, "Total general", font=FONT_BOLD, fill=FILL_HEADER)
    for ci, comp in enumerate(companies):
        val = mean_col(filter_rows(rows_sector, CompanyNameTo=comp), "Calification")
        if val is not None:
            set_cell(ws, row_total_general, ci + 2, round2(val), font=FONT_BOLD, number_format=NUMBER_FORMAT, fill=FILL_HEADER)
    val = mean_col(rows_sector, "Calification")
    if val is not None:
        set_cell(ws, row_total_general, num_companies + 2, round2(val), font=FONT_BOLD, number_format=NUMBER_FORMAT, fill=FILL_HEADER)

    current_row = row_total_general + 3

    # Seccion C: Resumen por Subcategoria
    ROW_RESUMEN_HEADER = current_row
    set_cell(ws, ROW_RESUMEN_HEADER, 1, "Subcategoria", font=FONT_HEADER, fill=FILL_HEADER)
    for ci, comp in enumerate(companies):
        set_cell(ws, ROW_RESUMEN_HEADER, ci + 2, comp, font=FONT_HEADER, fill=FILL_HEADER)
    set_cell(ws, ROW_RESUMEN_HEADER, num_companies + 2, "Total general", font=FONT_HEADER, fill=FILL_HEADER)

    current_row = ROW_RESUMEN_HEADER + 1
    company_subcat_vals = {comp: [] for comp in companies}
    present_subcats = ordered_subcategories(rows_sector)

    for subcat in present_subcats:
        set_cell(ws, current_row, 1, subcat, font=FONT_BASE)
        df_sub = filter_rows(rows_sector, PollDesc2=subcat)

        all_subcat_vals = []
        for ci, comp in enumerate(companies):
            val = mean_col(filter_rows(df_sub, CompanyNameTo=comp), "Calification")
            if val is not None:
                set_cell(ws, current_row, ci + 2, round2(val), number_format=NUMBER_FORMAT)
                company_subcat_vals[comp].append(val)
                all_subcat_vals.append(val)

        if all_subcat_vals:
            set_cell(ws, current_row, num_companies + 2, round2(sum(all_subcat_vals) / len(all_subcat_vals)), number_format=NUMBER_FORMAT)
        current_row += 1

    row_avg_subcat = current_row
    set_cell(ws, row_avg_subcat, 1, "Total general", font=FONT_BOLD, fill=FILL_HEADER)
    company_general_avg = {}
    for ci, comp in enumerate(companies):
        vals = company_subcat_vals[comp]
        if vals:
            avg = sum(vals) / len(vals)
            company_general_avg[comp] = round2(avg)
            set_cell(ws, row_avg_subcat, ci + 2, round2(avg), font=FONT_BOLD, number_format=NUMBER_FORMAT, fill=FILL_HEADER)

    all_avgs = list(company_general_avg.values())
    if all_avgs:
        set_cell(ws, row_avg_subcat, num_companies + 2, round2(sum(all_avgs) / len(all_avgs)), font=FONT_BOLD, number_format=NUMBER_FORMAT, fill=FILL_HEADER)

    current_row = row_avg_subcat + 3

    # Conteo de Evaluadores
    eval_dict = group_nunique(rows_sector, "CompanyNameTo", "CompanyNameFrom")

    # Seccion D: Ranking
    ROW_RANKING_TITLE = current_row
    set_cell(ws, ROW_RANKING_TITLE, 2, "EMPRESA", font=FONT_HEADER, fill=FILL_HEADER)
    set_cell(ws, ROW_RANKING_TITLE, 3, "PUNTAJE", font=FONT_HEADER, fill=FILL_HEADER)
    set_cell(ws, ROW_RANKING_TITLE, 4, "PUESTO", font=FONT_HEADER, fill=FILL_HEADER)
    set_cell(ws, ROW_RANKING_TITLE, 5, "CANT EVALUADORES", font=FONT_HEADER, fill=FILL_HEADER)

    ranking_data = sorted(company_general_avg.items(), key=lambda x: x[1], reverse=True)
    current_row = ROW_RANKING_TITLE + 1
    top_n = min(7, len(ranking_data))

    for rank, (comp, score) in enumerate(ranking_data, 1):
        if rank <= top_n:
            fill = FILL_GREEN
            font = Font(name="Aptos Narrow", size=11, bold=True, color="000000")
        else:
            fill = FILL_CELESTE
            font = Font(name="Aptos Narrow", size=11, color="FF0000")

        set_cell(ws, current_row, 2, comp, font=font, fill=fill)
        set_cell(ws, current_row, 3, score, font=font, fill=fill, number_format=NUMBER_FORMAT)
        set_cell(ws, current_row, 4, rank, font=font, fill=fill)
        set_cell(ws, current_row, 5, eval_dict.get(comp, 0), font=font, fill=fill)
        current_row += 1

    current_row += 2

    # Seccion E: Conteo de Evaluadores
    ROW_EVAL_TITLE = current_row
    set_cell(ws, ROW_EVAL_TITLE, 1, "Etiquetas de fila", font=FONT_HEADER, fill=FILL_HEADER)
    set_cell(ws, ROW_EVAL_TITLE, 2, "# Empresas que lo evaluaron", font=FONT_HEADER, fill=FILL_HEADER)

    current_row = ROW_EVAL_TITLE + 1
    for comp in sorted(eval_dict.keys()):
        set_cell(ws, current_row, 1, comp)
        set_cell(ws, current_row, 2, int(eval_dict[comp]))
        current_row += 1

    ws.column_dimensions["A"].width = 45
    for ci in range(2, total_cols + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 18


def generate_workbook(rows, output_path):
    """Construye el workbook a partir de una lista de records (dicts) y lo guarda."""
    wb = Workbook()
    wb.remove(wb.active)

    for id_category, sector_name in SECTOR_MAP.items():
        for id_type, type_short in TYPE_MAP.items():
            df_filtered = filter_rows(rows, IDCategory=id_category, IDType=id_type)
            id_poll = df_filtered[0]["IDPoll"] if df_filtered else "DV"
            build_sheet(wb, df_filtered, id_poll, id_category, id_type, sector_name, type_short)

    wb.save(output_path)
    return output_path


def main():
    print("Leyendo archivo fuente...")
    rows = read_xlsx_records(find_input_file())
    auto_eval = sum(1 for r in rows if r.get("CompanyNameFrom") == r.get("CompanyNameTo"))
    rows = [r for r in rows if r.get("CompanyNameFrom") != r.get("CompanyNameTo")]
    print(f"  Filas totales: {len(rows)} (excluidas {auto_eval} auto-evaluaciones)")
    print(f"\nGuardando resultado en: {OUTPUT_FILE}")
    generate_workbook(rows, OUTPUT_FILE)
    print("Listo!")


if __name__ == "__main__":
    main()
